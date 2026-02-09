import sys
from datetime import datetime, timedelta
from openpyxl import load_workbook

# === CONFIGURATION ===
STUDENT_SHEET = ["Agile", "Narzędzia"]
DATE_SHEET_SUFFIX = "-daty"
TITLE_SHEET_SUFFIX = "-tematy"

# Columns
EMAIL_COL_MAIN = 3  # D in 0-based index

NAME_COL = 0        # A
SURNAME_COL = 1     # B
ENGLISH_COL = 5     # F -> "Yes" if English lecture

# Columns for Date and Title sheets
EMAIL_COL = 2       # C
VALUE_COL = 4       # E

# Schedule
START_DATE = datetime(2025, 10, 9)
WEEK_INTERVAL = 7
DURATION_MONTHS = 4
# ======================

def load_lookup(sheet):
    """Return dict: email -> value (date string or title string)."""
    lookup = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if len(row) <= max(EMAIL_COL, VALUE_COL):
            continue
        email = str(row[EMAIL_COL]).strip()
        val = row[VALUE_COL]
        if email:
            lookup[email] = val
    return lookup

def parse_date(value):
    """Try to convert string or Excel date to datetime.date."""
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, str):
        for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%Y/%m/%d", "%d/%m/%Y"):
            try:
                return datetime.strptime(value.strip(), fmt).date()
            except ValueError:
                continue
    return None

def generate_weekly_schedule(wb, sheet_name):
    student_sheet = wb[sheet_name]
    date_sheet = wb[sheet_name + DATE_SHEET_SUFFIX]
    title_sheet = wb[sheet_name + TITLE_SHEET_SUFFIX]

    # Load lookups
    date_lookup = load_lookup(date_sheet)
    title_lookup = load_lookup(title_sheet)


    # Build lectures list
    lectures = []
    for row in student_sheet.iter_rows(min_row=2, values_only=True):
        if len(row) <= EMAIL_COL_MAIN:
            continue
        email = str(row[EMAIL_COL_MAIN]).strip()
        if not email:
            continue
        name = row[NAME_COL] if len(row) > NAME_COL else ""
        surname = row[SURNAME_COL] if len(row) > SURNAME_COL else ""
        english = (str(row[ENGLISH_COL]).strip().lower() == "yes") if len(row) > ENGLISH_COL else False

        lecture_date = parse_date(date_lookup.get(email))
        lecture_title = title_lookup.get(email, "")

        if lecture_date and lecture_title:
            lectures.append({
                "date": lecture_date,
                "title": lecture_title.strip(),
                "name": f"{name} {surname}".strip(),
                "english": english
            })

    lectures.sort(key=lambda x: x["date"])

    # Generate Markdown schedule week-by-week
    md = ""
    current_date = START_DATE.date()
    end_date = (START_DATE + timedelta(days=30*DURATION_MONTHS)).date()
    week_number = 1

    while current_date <= end_date:
        display_date = current_date.strftime("%d.%m.%Y")
        # Lectures on this date
        week_lectures = [lec for lec in lectures if lec["date"] == current_date]

        if not week_lectures:
            md += f"{display_date} - **Wykład nie odbędzie się**\n\n"
        else:
            for lec in week_lectures:
                if lec["english"]:
                    md += f"{display_date} – Lecture {week_number}: **{lec['title']}** – *{lec['name']}* (open lecture) *Lecture will be held in English*\n\n"
                else:
                    md += f"{display_date} – Wykład {week_number}: **{lec['title']}** – *{lec['name']}* (wykład otwarty)\n\n"
                week_number += 1

        current_date += timedelta(days=WEEK_INTERVAL)

    return md

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python weekly_schedule_from_lookup.py <excel_file.xlsx>")
        sys.exit(1)

    file_path = sys.argv[1]
    wb = load_workbook(filename=file_path)

    for sheet_name in STUDENT_SHEET:
        md = generate_weekly_schedule(wb, sheet_name)

        output_file = f"{sheet_name}_weekly_schedule.txt"
        with open(output_file, "w", encoding="utf-8") as f:
            f.write(md)
        print(f"Saved schedule: {output_file}")
