import sys
from datetime import datetime, date
from openpyxl import load_workbook

# === CONFIGURATION ===
NR_INDEKSU_COL = 2       # Column index for "Nr Indeksu" (0-based)
SHEETS = ["Agile", "Narzędzia"]  # Sheet names to process
# ======================


def looks_like_date(value: str) -> bool:
    """Return True if the string looks like a date in common formats."""
    if not isinstance(value, str):
        return False
    value = value.strip()
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%Y/%m/%d", "%d/%m/%Y"):
        try:
            datetime.strptime(value, fmt)
            return True
        except ValueError:
            continue
    return False


def is_date_like(value) -> bool:
    """Return True if the cell value is a datetime/date or looks like a date string."""
    if isinstance(value, (datetime, date)):
        return True
    if isinstance(value, str) and looks_like_date(value):
        return True
    return False


def generate_markdown_table(file_path, sheet_name):
    workbook = load_workbook(filename=file_path)

    if sheet_name not in workbook.sheetnames:
        print(f"Sheet '{sheet_name}' not found in the Excel file.")
        sys.exit(1)

    sheet = workbook[sheet_name]
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))

    # Dynamically detect lecture columns — only date-like headers
    lecture_cols = [
        i for i, val in enumerate(header_row)
        if i > NR_INDEKSU_COL and is_date_like(val)
    ]

    if not lecture_cols:
        print(f"No date columns found in sheet '{sheet_name}'.")
        return ""

    # Collect data rows
    rows = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        nr_indeksu = row[NR_INDEKSU_COL] if len(row) > NR_INDEKSU_COL else None
        if not nr_indeksu:
            continue
        rows.append(row)

    # Sort by Nr Indeksu (converted to string)
    rows.sort(key=lambda r: str(r[NR_INDEKSU_COL]).strip())

    # --- Generate Markdown ---
    header_cells = ["Nr indeksu"]
    for i in lecture_cols:
        val = header_row[i]
        if isinstance(val, (datetime, date)):
            val = val.strftime("%Y-%m-%d")
        header_cells.append(str(val))

    md = "<style>\n\ttable, th, td {\n\t\tborder: 1px solid black;\n\t\tborder-collapse: collapse;\n\t}\n</style>\n\n"

    md += "| " + " | ".join(header_cells) + " |\n"
    md += "| " + " | ".join(["---"] * len(header_cells)) + " |\n"

    for row in rows:
        cells = [str(row[NR_INDEKSU_COL])]
        for col in lecture_cols:
            value = row[col] if len(row) > col else None
            cells.append("+" if value else "")
        md += "| " + " | ".join(cells) + " |\n"

    return md


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python table_generator.py <excel_file.xlsx>")
        sys.exit(1)

    file_path = sys.argv[1]

    for sheet_name in SHEETS:
        markdown = generate_markdown_table(file_path, sheet_name)
        if markdown:
            output_file = f"{sheet_name}.txt"
            with open(output_file, "w", encoding="utf-8") as f:
                f.write(markdown)
            print(f"Saved: {output_file}")
